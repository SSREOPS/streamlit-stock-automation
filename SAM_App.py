import yfinance as yf # type: ignore
import streamlit as st # type: ignore
import pandas as pd # type: ignore
import io
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment # type: ignore


# Function to get stock metrics for a single ticker
def Get_Stock_Data(stock_ticker):
    # Fetch data from Yahoo Finance
    ticker = yf.Ticker(stock_ticker)
    stats = ticker.info  # Dictionary with all the stock metrics

    # Define which metrics to extract
    keys = [
        'currentPrice', 'marketCap', 'ebitda', 'trailingPE', 'totalRevenue',
        'fiftyTwoWeekHigh', 'fiftyTwoWeekLow', 'trailingEps', 'totalDebt',
        'totalCash', 'operatingCashflow', 'freeCashflow'
    ]

    # Human-readable labels for each metric
    internal_key_map = {
        'currentPrice': 'Current Price',
        'marketCap': 'Market cap (B)',
        'ebitda': 'EBITDA (B)',
        'trailingPE': 'PE',
        'totalRevenue': 'Revenue (TTM) (B)',
        'fiftyTwoWeekHigh': '52 Week H',
        'fiftyTwoWeekLow': '52 Week L',
        'trailingEps': 'EPS',
        'totalDebt': 'Total debt (B)',
        'totalCash': 'Cash Reserve (B)',
        'operatingCashflow': 'Operating cashflow (TTM) (B)',
        'freeCashflow': 'Levered Free Cash Flow (TTM) (B)',
    }

    # Create a list of tuples: (Yahoo key, Human-readable name, Raw value)
    data = [(key, internal_key_map.get(key), stats.get(key)) for key in keys]

    # Build the DataFrame
    # Note: we're using 'Ticker' as the column name for internal keys for easier transposing later
    df = pd.DataFrame(data, columns=['Actual Metric', 'Ticker', 'Raw Value'])

    # Format Pandas to show full float values instead of scientific notation
    pd.set_option('display.float_format', '{:.2f}'.format)

    # Add a new column named after the stock ticker symbol.
    # If the metric includes '(B)', we assume it's in billions and divide by 1e9.
    df[stock_ticker] = df.apply(
        lambda row: round(row['Raw Value'] / 1e9, 2) if '(B)' in row['Ticker'] and isinstance(row['Raw Value'], (int, float)) else row['Raw Value'],
        axis=1
    )

    # Keep only the internal key names and the transformed values
    df = df.iloc[:, [1, 3]]  # Columns: 'Ticker' (internal key), and the stock value

    # Set internal key as the index
    df.set_index('Ticker', inplace=True)

    # Transpose the DataFrame so the ticker becomes the row label
    trans_df = df.T

    # Format all numeric values to 2 decimal places as strings
    #trans_df = trans_df.map(lambda x: f"{x:.2f}" if isinstance(x, (int, float)) else x)

    # Return the final, cleaned transposed DataFrame (1 row, many columns)
    return trans_df


# Function to get stock data for multiple tickers and combine into one DataFrame
def Get_Stock_Data_Multi(stock_ticker_list):
    final_df = pd.DataFrame()
    #
    for stock_ticker in stock_ticker_list:
        try:
            df = Get_Stock_Data(stock_ticker)
            final_df = pd.concat([final_df, df], ignore_index=False)
        except Exception as e:
            print(f"⚠️ Error fetching data for {stock_ticker}: {e}")
    #
    return final_df


# Processing Excel for download
def to_excel_buffer(df: pd.DataFrame) -> io.BytesIO:
    buffer = io.BytesIO()
    df_to_download = df
    df_to_download.index.name = 'Ticker'
    #
    # Use ExcelWriter to write and format Excel
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_to_download.to_excel(writer, sheet_name='StockData')
        #
        workbook = writer.book
        worksheet = writer.sheets['StockData']
        #
        # Define styles
        header_font = Font(name='Aptos narrow', bold=True, size=14)
        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow fill
        #
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        #
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.row_dimensions[1].height = 90  # adjust as needed
        #
        # Format header row (row 1)
        for col_idx, col in enumerate(df.columns, 2):  # +2 because index is col 1
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        #
        # Format index header (A1)
        index_header_cell = worksheet.cell(row=1, column=1)
        index_header_cell.font = header_font
        index_header_cell.fill = header_fill
        index_header_cell.alignment = header_alignment
        #
        # Format all data cells with border and alignment
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
                cell.font = Font(size=14)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
        #
        # Set column widths
        column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
        column_widths  = [70, 85, 75, 85, 70, 95, 70, 70, 70, 70, 85, 100, 135]
        #
        for col_letter, width in zip(column_letters, column_widths):
            worksheet.column_dimensions[col_letter].width = width/7
    #       
    buffer.seek(0)
    return buffer


# Streamlit App
def main():
    # Hide Streamlit header and footer
    hide_streamlit_style = '''
        <style>
        #MainMenu {visibility: hidden;}
        header {visibility: hidden;}
        footer {visibility: hidden;}
        </style>
        '''
    st.markdown(hide_streamlit_style, unsafe_allow_html=True)
    st.set_page_config(page_title='Live Stock Data', layout='wide')

    # 1. Create UI to input ticker symbols
    st.title('Live Stock Statistics')

    tickers_input = st.text_input("Enter stock tickers separated by commas", "TGT, LULU, WEN")

    # 2. Button to trigger data fetch
    if st.button("Get Stock Data"):
        # 3. Parse input into a list, strip whitespace
        ticker_list = [ticker.strip().upper() for ticker in tickers_input.split(",") if ticker.strip()]

        if ticker_list:
            # 4. Call your Get_Stock_Data_Multi function here
            df = Get_Stock_Data_Multi(ticker_list)

            if not df.empty:
                # 5. Display dataframe on screen
                rounded_df = df.map(lambda x: f"{x:.2f}" if isinstance(x, (int, float)) else x)

                # Display the DataFrame with some styling
                st.markdown('### Statistics:')

                st.table(rounded_df)

                st.download_button(
                    label="Download data as Excel",
                    data=to_excel_buffer(df),
                    file_name='stock_data.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            else:
                st.warning("No data found for the given tickers.")
        else:
            st.error("Please enter at least one ticker.")

if __name__ == '__main__':

    main()



